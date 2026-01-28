"""Excel reader with LLM-powered column detection.

Reads an Excel file, uses OpenAI to identify which columns contain
first name, last name, and company, then loads rows into the database.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl
from openai import OpenAI

from ..db.repository import Repository
from ..models import ColumnMapping, PersonInput

logger = logging.getLogger(__name__)

_COLUMN_DETECTION_PROMPT = """You are analyzing an Excel spreadsheet. Given the column headers and a few sample rows, identify which columns contain:
1. First name (Vorname / first_name / ad / isim / Name)
2. Last name (Nachname / last_name / soyad / surname / Familienname)
3. Company name (Firma / company / Unternehmen / şirket / sirket / organization)

Return a JSON object with exactly these keys:
- "first_name_col": the exact column header text for first name
- "last_name_col": the exact column header text for last name
- "company_col": the exact column header text for company name

If a single column contains the full name (first + last), set both first_name_col and last_name_col to that column header and add "full_name": true.

Return ONLY the JSON object, no other text.

Column headers: {headers}

Sample data (first 3 rows):
{samples}"""


def detect_columns(
    headers: list[str],
    sample_rows: list[list[str]],
    openai_api_key: str,
    model: str = "gpt-4o-mini",
) -> ColumnMapping:
    """Use OpenAI to detect which columns map to first_name, last_name, company."""
    samples_text = "\n".join(
        [", ".join(str(cell) for cell in row) for row in sample_rows[:3]]
    )

    client = OpenAI(api_key=openai_api_key)
    response = client.chat.completions.create(
        model=model,
        messages=[
            {
                "role": "user",
                "content": _COLUMN_DETECTION_PROMPT.format(
                    headers=headers,
                    samples=samples_text,
                ),
            }
        ],
        response_format={"type": "json_object"},
        temperature=0,
    )

    content = response.choices[0].message.content or "{}"
    result = json.loads(content)
    logger.info("Detected column mapping: %s", result)

    return ColumnMapping(
        first_name_col=result["first_name_col"],
        last_name_col=result["last_name_col"],
        company_col=result["company_col"],
    )


def read_excel_headers_and_samples(
    file_path: Path,
) -> tuple[list[str], list[list[str]]]:
    """Read column headers and first 5 sample rows from an Excel file."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in {file_path}")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError(f"Empty Excel file: {file_path}")

    headers = [str(cell) if cell is not None else "" for cell in header_row]
    samples = []
    for i, row in enumerate(rows_iter):
        if i >= 5:
            break
        samples.append([str(cell) if cell is not None else "" for cell in row])

    wb.close()
    return headers, samples


def load_excel_to_db(
    file_path: Path,
    repo: Repository,
    column_mapping: ColumnMapping,
) -> int:
    """Load all person rows from Excel into the database.

    Returns the number of rows loaded. Idempotent — existing rows are skipped.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in {file_path}")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError(f"Empty Excel file: {file_path}")

    headers = [str(cell) if cell is not None else "" for cell in header_row]

    # Build column index map
    col_map = {name: idx for idx, name in enumerate(headers)}
    fn_idx = col_map.get(column_mapping.first_name_col)
    ln_idx = col_map.get(column_mapping.last_name_col)
    co_idx = col_map.get(column_mapping.company_col)

    if fn_idx is None:
        raise ValueError(f"Column '{column_mapping.first_name_col}' not found in Excel headers")
    if ln_idx is None:
        raise ValueError(f"Column '{column_mapping.last_name_col}' not found in Excel headers")
    if co_idx is None:
        raise ValueError(f"Column '{column_mapping.company_col}' not found in Excel headers")

    bulk_rows: list[tuple[int, str, str, str]] = []
    row_count = 0

    for row_idx, row in enumerate(rows_iter, start=1):
        first_name = str(row[fn_idx]).strip() if row[fn_idx] else ""
        last_name = str(row[ln_idx]).strip() if row[ln_idx] else ""
        company = str(row[co_idx]).strip() if row[co_idx] else ""

        if not first_name or not last_name:
            logger.warning("Skipping row %d: missing name (first='%s', last='%s')",
                           row_idx, first_name, last_name)
            continue

        bulk_rows.append((row_idx, first_name, last_name, company))
        row_count += 1

        # Flush in chunks of 500
        if len(bulk_rows) >= 500:
            repo.upsert_rows_bulk(bulk_rows)
            bulk_rows.clear()

    # Flush remaining
    if bulk_rows:
        repo.upsert_rows_bulk(bulk_rows)

    wb.close()
    logger.info("Loaded %d rows from %s", row_count, file_path)
    return row_count


def get_persons_from_db(repo: Repository, api: str, status: str) -> list[PersonInput]:
    """Fetch pending persons from DB as PersonInput models."""
    rows = repo.get_rows_by_status(api, status)
    return [
        PersonInput(
            row_id=r.row_id,
            first_name=r.first_name,
            last_name=r.last_name,
            company=r.company,
        )
        for r in rows
    ]
