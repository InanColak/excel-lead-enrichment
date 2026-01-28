"""Excel writer that adds enrichment columns to the original file."""

from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from ..db.repository import Repository

logger = logging.getLogger(__name__)

# Output column names in the order they will be appended
OUTPUT_COLUMNS = [
    "apollo_email",
    "apollo_handynummer",
    "apollo_festnetz_durchwahl",
    "lusha_email",
    "lusha_handynummer",
    "lusha_festnetz_durchwahl",
    "Email",
    "Telefonnummer",
    "Durchwahl_festnetz",
]

# Cell colors
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")


def _normalize_phone(phone: str | None) -> str:
    """Normalize a phone number for comparison.

    Removes spaces, dashes, parentheses, dots.
    Keeps + at the beginning and digits only.
    """
    if not phone:
        return ""
    # Keep only + (at start) and digits
    normalized = re.sub(r"[^\d+]", "", phone)
    return normalized


def _is_zentrale(phone: str | None) -> bool:
    """Check if a phone number is a Zentrale (switchboard) number.

    Zentrale numbers typically end with -0 or just 0 as the extension.
    Examples: +49 721 25516-0, +49 721 255160
    """
    if not phone:
        return False
    # Check if original has -0 pattern
    if re.search(r"-0\s*$", phone):
        return True
    return False


def _compare_values(apollo_val: str | None, lusha_val: str | None) -> tuple[str, PatternFill | None]:
    """Compare Apollo and Lusha values and return (text, fill_color).

    Returns:
        - ("gleich", GREEN_FILL) if both have same non-empty value
        - ("ungleich", ORANGE_FILL) if both have different non-empty values
        - ("Apollo", None) if only Apollo has a value
        - ("Lusha", None) if only Lusha has a value
        - ("", None) if both are empty
    """
    apollo_has = bool(apollo_val and apollo_val.strip())
    lusha_has = bool(lusha_val and lusha_val.strip())

    if apollo_has and lusha_has:
        # Normalize for comparison (strip whitespace, lowercase)
        if apollo_val.strip().lower() == lusha_val.strip().lower():
            return "gleich", GREEN_FILL
        else:
            return "ungleich", ORANGE_FILL
    elif apollo_has:
        return "Apollo", None
    elif lusha_has:
        return "Lusha", None
    else:
        return "", None


def _compare_phones(apollo_val: str | None, lusha_val: str | None) -> tuple[str, PatternFill | None]:
    """Compare phone numbers with normalization.

    Normalizes phone numbers before comparison to handle format differences.
    """
    apollo_has = bool(apollo_val and apollo_val.strip())
    lusha_has = bool(lusha_val and lusha_val.strip())

    if apollo_has and lusha_has:
        # Normalize both numbers
        apollo_norm = _normalize_phone(apollo_val)
        lusha_norm = _normalize_phone(lusha_val)

        if apollo_norm == lusha_norm:
            return "gleich", GREEN_FILL
        else:
            return "ungleich", ORANGE_FILL
    elif apollo_has:
        return "Apollo", None
    elif lusha_has:
        return "Lusha", None
    else:
        return "", None


def _compare_durchwahl(apollo_val: str | None, lusha_val: str | None) -> tuple[str, PatternFill | None]:
    """Compare Durchwahl/Festnetz numbers with Zentrale detection.

    If one is Zentrale (-0) and other is a real Durchwahl, they are different.
    Otherwise, compare normalized numbers.
    """
    apollo_has = bool(apollo_val and apollo_val.strip())
    lusha_has = bool(lusha_val and lusha_val.strip())

    if apollo_has and lusha_has:
        apollo_zentrale = _is_zentrale(apollo_val)
        lusha_zentrale = _is_zentrale(lusha_val)

        # If one is Zentrale and other is not, they're different
        if apollo_zentrale != lusha_zentrale:
            return "ungleich", ORANGE_FILL

        # Both are same type (both Zentrale or both Durchwahl), compare normalized
        apollo_norm = _normalize_phone(apollo_val)
        lusha_norm = _normalize_phone(lusha_val)

        if apollo_norm == lusha_norm:
            return "gleich", GREEN_FILL
        else:
            return "ungleich", ORANGE_FILL
    elif apollo_has:
        return "Apollo", None
    elif lusha_has:
        return "Lusha", None
    else:
        return "", None


def write_enriched_excel(
    input_path: Path,
    output_path: Path,
    repo: Repository,
) -> Path:
    """Open the original Excel, append enrichment columns, save to output_path.

    Preserves all original data and formatting. Returns the output path.
    """
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in {input_path}")

    last_col = ws.max_column

    # Write new column headers in row 1
    for offset, col_name in enumerate(OUTPUT_COLUMNS):
        ws.cell(row=1, column=last_col + 1 + offset, value=col_name)

    # Fetch all enriched rows from DB
    all_rows = repo.get_all_rows()

    for row in all_rows:
        excel_row = row.row_id + 1  # row_id is 1-indexed data row, Excel row 1 is headers
        base_col = last_col + 1

        # Write Apollo columns
        ws.cell(row=excel_row, column=base_col + 0, value=row.apollo_email)
        ws.cell(row=excel_row, column=base_col + 1, value=row.apollo_mobile)
        ws.cell(row=excel_row, column=base_col + 2, value=row.apollo_direct)

        # Write Lusha columns
        ws.cell(row=excel_row, column=base_col + 3, value=row.lusha_email)
        ws.cell(row=excel_row, column=base_col + 4, value=row.lusha_mobile)
        ws.cell(row=excel_row, column=base_col + 5, value=row.lusha_direct)

        # Write comparison columns
        # Email comparison
        email_text, email_fill = _compare_values(row.apollo_email, row.lusha_email)
        cell_email = ws.cell(row=excel_row, column=base_col + 6, value=email_text)
        if email_fill:
            cell_email.fill = email_fill

        # Telefonnummer (mobile) comparison - with phone normalization
        phone_text, phone_fill = _compare_phones(row.apollo_mobile, row.lusha_mobile)
        cell_phone = ws.cell(row=excel_row, column=base_col + 7, value=phone_text)
        if phone_fill:
            cell_phone.fill = phone_fill

        # Durchwahl/Festnetz (direct) comparison - with Zentrale detection
        direct_text, direct_fill = _compare_durchwahl(row.apollo_direct, row.lusha_direct)
        cell_direct = ws.cell(row=excel_row, column=base_col + 8, value=direct_text)
        if direct_fill:
            cell_direct.fill = direct_fill

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    logger.info("Enriched Excel written to %s (%d rows)", output_path, len(all_rows))
    return output_path
