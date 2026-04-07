"""OUTPUT-01, D-55, D-56, D-57, D-64: Enriched Excel output generation tests.

Tests the output generation module: status mapping logic and end-to-end
Excel output generation with correct column appending and row alignment.
"""

import uuid
from contextlib import asynccontextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.contacts.models import Contact
from app.jobs.models import Job, JobRow, RowStatus


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _mock_session_factory(test_session):
    """Build a mock session factory that yields the test session."""

    class _Factory:
        def __call__(self):
            @asynccontextmanager
            async def _ctx():
                yield test_session

            return _ctx()

    return _Factory()


def _create_test_xlsx(path: Path) -> Path:
    """Create a simple test xlsx with 3 data rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(["First Name", "Last Name", "Email", "Company"])
    ws.append(["John", "Doe", "john@example.com", "Acme Corp"])
    ws.append(["Jane", "Smith", "jane@smith.co", "Globex Inc"])
    ws.append(["Bob", "Wilson", "bob@wilson.io", "Initech"])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# map_enrichment_status tests
# ---------------------------------------------------------------------------


class TestMapEnrichmentStatus:
    """D-55: Status mapping for output generation."""

    def test_enriched_with_phone(self):
        from app.jobs.output import map_enrichment_status

        contact = Contact(email="a@b.com", phone="+15551234567")
        assert map_enrichment_status("enriched", contact) == "enriched"

    def test_enriched_without_phone(self):
        from app.jobs.output import map_enrichment_status

        contact = Contact(email="a@b.com", phone=None)
        assert map_enrichment_status("enriched", contact) == "email_only"

    def test_enriched_empty_phone(self):
        from app.jobs.output import map_enrichment_status

        contact = Contact(email="a@b.com", phone="")
        assert map_enrichment_status("enriched", contact) == "email_only"

    def test_email_only_status(self):
        from app.jobs.output import map_enrichment_status

        contact = Contact(email="a@b.com", phone="+15551234567")
        assert map_enrichment_status("email_only", contact) == "email_only"

    def test_not_found(self):
        from app.jobs.output import map_enrichment_status

        assert map_enrichment_status("not_found", None) == "not_found"

    def test_error(self):
        from app.jobs.output import map_enrichment_status

        assert map_enrichment_status("error", None) == "error"

    def test_skipped_maps_to_error(self):
        from app.jobs.output import map_enrichment_status

        assert map_enrichment_status("skipped", None) == "error"

    def test_unknown_status_fallback(self):
        from app.jobs.output import map_enrichment_status

        assert map_enrichment_status("unknown_thing", None) == "pending"


# ---------------------------------------------------------------------------
# generate_output_file tests
# ---------------------------------------------------------------------------


class TestGenerateOutputFile:
    """D-56, D-57, D-64: End-to-end output file generation."""

    async def test_generates_xlsx_with_appended_columns(
        self, test_session, regular_user, tmp_path
    ):
        """Output xlsx has original columns + enriched_email, enriched_phone, enrichment_status."""
        from app.jobs.output import generate_output_file

        xlsx_path = _create_test_xlsx(tmp_path / "original.xlsx")

        # Create job
        job = Job(
            id=uuid.uuid4(),
            user_id=regular_user.id,
            filename="contacts.xlsx",
            file_path=str(xlsx_path),
            status="complete",
            total_rows=3,
            valid_rows=3,
            error_rows=0,
        )
        test_session.add(job)
        await test_session.flush()

        # Create contacts
        contact1 = Contact(email="john@enriched.com", phone="+15551111111")
        contact2 = Contact(email="jane@enriched.com", phone=None)
        test_session.add(contact1)
        test_session.add(contact2)
        await test_session.flush()

        # Create job rows
        row0 = JobRow(
            id=uuid.uuid4(),
            job_id=job.id,
            row_index=0,
            raw_data={"First Name": "John", "Last Name": "Doe"},
            status="enriched",
            contact_id=contact1.id,
        )
        row1 = JobRow(
            id=uuid.uuid4(),
            job_id=job.id,
            row_index=1,
            raw_data={"First Name": "Jane", "Last Name": "Smith"},
            status="enriched",
            contact_id=contact2.id,
        )
        row2 = JobRow(
            id=uuid.uuid4(),
            job_id=job.id,
            row_index=2,
            raw_data={"First Name": "Bob", "Last Name": "Wilson"},
            status="not_found",
            contact_id=None,
        )
        test_session.add(row0)
        test_session.add(row1)
        test_session.add(row2)
        await test_session.flush()

        factory = _mock_session_factory(test_session)
        output_path = await generate_output_file(job.id, factory)

        # Verify output file exists
        assert Path(output_path).exists()

        # Verify output filename pattern
        assert output_path.endswith("_enriched.xlsx")
        assert "contacts_enriched.xlsx" in output_path

        # Load and verify content
        wb = load_workbook(output_path)
        ws = wb.active

        # Header row should have original 4 + 3 appended = 7 columns
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "First Name",
            "Last Name",
            "Email",
            "Company",
            "enriched_email",
            "enriched_phone",
            "enrichment_status",
        ]

        # Row 2 (row_index 0): John - enriched with phone
        row2_vals = [cell.value for cell in ws[2]]
        assert row2_vals[0] == "John"
        assert row2_vals[1] == "Doe"
        assert row2_vals[4] == "john@enriched.com"
        assert row2_vals[5] == "+15551111111"
        assert row2_vals[6] == "enriched"

        # Row 3 (row_index 1): Jane - enriched without phone (email_only)
        row3_vals = [cell.value for cell in ws[3]]
        assert row3_vals[0] == "Jane"
        assert row3_vals[4] == "jane@enriched.com"
        assert row3_vals[5] is None
        assert row3_vals[6] == "email_only"

        # Row 4 (row_index 2): Bob - not_found
        row4_vals = [cell.value for cell in ws[4]]
        assert row4_vals[0] == "Bob"
        assert row4_vals[4] is None
        assert row4_vals[5] is None
        assert row4_vals[6] == "not_found"

        wb.close()

    async def test_preserves_original_data(self, test_session, regular_user, tmp_path):
        """Original cell values are preserved exactly in output."""
        from app.jobs.output import generate_output_file

        xlsx_path = _create_test_xlsx(tmp_path / "original.xlsx")

        job = Job(
            id=uuid.uuid4(),
            user_id=regular_user.id,
            filename="data.xlsx",
            file_path=str(xlsx_path),
            status="complete",
            total_rows=3,
            valid_rows=3,
            error_rows=0,
        )
        test_session.add(job)
        await test_session.flush()

        # Create rows with no contacts (all not_found)
        for idx in range(3):
            row = JobRow(
                id=uuid.uuid4(),
                job_id=job.id,
                row_index=idx,
                raw_data={"First Name": f"Name{idx}"},
                status="not_found",
                contact_id=None,
            )
            test_session.add(row)
        await test_session.flush()

        factory = _mock_session_factory(test_session)
        output_path = await generate_output_file(job.id, factory)

        wb_out = load_workbook(output_path)
        ws_out = wb_out.active

        # Original data rows preserved
        assert ws_out[2][0].value == "John"
        assert ws_out[2][1].value == "Doe"
        assert ws_out[2][2].value == "john@example.com"
        assert ws_out[2][3].value == "Acme Corp"

        assert ws_out[3][0].value == "Jane"
        assert ws_out[3][1].value == "Smith"

        assert ws_out[4][0].value == "Bob"
        assert ws_out[4][1].value == "Wilson"

        wb_out.close()

    async def test_row_index_alignment(self, test_session, regular_user, tmp_path):
        """row_index 0 maps to Excel row 2 (row 1 is header) -- off-by-one prevention."""
        from app.jobs.output import generate_output_file

        xlsx_path = _create_test_xlsx(tmp_path / "original.xlsx")

        job = Job(
            id=uuid.uuid4(),
            user_id=regular_user.id,
            filename="align.xlsx",
            file_path=str(xlsx_path),
            status="complete",
            total_rows=3,
            valid_rows=3,
            error_rows=0,
        )
        test_session.add(job)
        await test_session.flush()

        contact = Contact(email="specific@test.com", phone="+10000000000")
        test_session.add(contact)
        await test_session.flush()

        # Only row_index=1 has enrichment (Jane)
        for idx in range(3):
            row = JobRow(
                id=uuid.uuid4(),
                job_id=job.id,
                row_index=idx,
                raw_data={},
                status="enriched" if idx == 1 else "not_found",
                contact_id=contact.id if idx == 1 else None,
            )
            test_session.add(row)
        await test_session.flush()

        factory = _mock_session_factory(test_session)
        output_path = await generate_output_file(job.id, factory)

        wb = load_workbook(output_path)
        ws = wb.active

        # row_index=1 -> Excel row 3 (1-based, row 1=header, row 2=index0, row 3=index1)
        row3_vals = [cell.value for cell in ws[3]]
        assert row3_vals[-3] == "specific@test.com"  # enriched_email
        assert row3_vals[-2] == "+10000000000"  # enriched_phone
        assert row3_vals[-1] == "enriched"  # enrichment_status

        # row_index=0 -> Excel row 2, should be not_found
        row2_vals = [cell.value for cell in ws[2]]
        assert row2_vals[-1] == "not_found"

        wb.close()

    async def test_stores_output_path_on_job(
        self, test_session, regular_user, tmp_path
    ):
        """Job.output_file_path is set after generation."""
        from app.jobs.output import generate_output_file

        xlsx_path = _create_test_xlsx(tmp_path / "original.xlsx")

        job = Job(
            id=uuid.uuid4(),
            user_id=regular_user.id,
            filename="stored.xlsx",
            file_path=str(xlsx_path),
            status="complete",
            total_rows=3,
            valid_rows=3,
            error_rows=0,
        )
        test_session.add(job)
        await test_session.flush()

        for idx in range(3):
            row = JobRow(
                id=uuid.uuid4(),
                job_id=job.id,
                row_index=idx,
                raw_data={},
                status="not_found",
            )
            test_session.add(row)
        await test_session.flush()

        factory = _mock_session_factory(test_session)
        output_path = await generate_output_file(job.id, factory)

        await test_session.refresh(job)
        assert job.output_file_path == output_path
        assert "stored_enriched.xlsx" in job.output_file_path

    async def test_output_filename_pattern(
        self, test_session, regular_user, tmp_path
    ):
        """D-57: Output filename is {stem}_enriched.xlsx."""
        from app.jobs.output import generate_output_file

        xlsx_path = _create_test_xlsx(tmp_path / "original.xlsx")

        job = Job(
            id=uuid.uuid4(),
            user_id=regular_user.id,
            filename="my_contacts_2024.xlsx",
            file_path=str(xlsx_path),
            status="complete",
            total_rows=3,
            valid_rows=3,
            error_rows=0,
        )
        test_session.add(job)
        await test_session.flush()

        for idx in range(3):
            row = JobRow(
                id=uuid.uuid4(),
                job_id=job.id,
                row_index=idx,
                raw_data={},
                status="not_found",
            )
            test_session.add(row)
        await test_session.flush()

        factory = _mock_session_factory(test_session)
        output_path = await generate_output_file(job.id, factory)

        assert Path(output_path).name == "my_contacts_2024_enriched.xlsx"
