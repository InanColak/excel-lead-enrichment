import uuid
from pathlib import Path
from unittest.mock import AsyncMock

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.auth.models import User
from app.auth.service import create_access_token, hash_password
from app.config import settings
from app.deps import get_db, get_redis
from app.main import app
from app.models.base import Base

# Import all models so Base.metadata knows about them for create_all/drop_all
from app.admin.models import ApiConfig  # noqa: F401
from app.contacts.models import Contact  # noqa: F401
from app.jobs.models import Job, JobRow  # noqa: F401

# Use the real PostgreSQL database from settings (reads DATABASE_URL from environment).
# Tests MUST run inside Docker: docker compose exec api pytest tests/ -v
# SQLite is NOT used -- CLAUDE.md forbids it (no JSONB, no UUID, no concurrent writes).


@pytest.fixture
async def test_session():
    """Per-test session with transaction rollback for isolation.

    Creates a fresh engine per test to avoid event-loop mismatches with asyncpg.
    Tables already exist from Alembic migrations run at container startup.
    """
    engine = create_async_engine(settings.database_url, echo=False)
    async with engine.connect() as conn:
        trans = await conn.begin()
        session = AsyncSession(bind=conn, expire_on_commit=False)
        yield session
        await session.close()
        await trans.rollback()
    await engine.dispose()


@pytest.fixture
def mock_redis():
    """Mock Redis client. Tests do not need a real Redis connection."""
    redis = AsyncMock()
    redis.exists = AsyncMock(return_value=0)  # No tokens revoked by default
    redis.ping = AsyncMock(return_value=True)
    redis.setex = AsyncMock()
    redis.sadd = AsyncMock()
    redis.smembers = AsyncMock(return_value=set())
    redis.delete = AsyncMock()
    return redis


@pytest.fixture
async def async_client(test_session, mock_redis):
    """httpx AsyncClient with dependency overrides for DB and Redis."""

    async def override_get_db():
        yield test_session

    async def override_get_redis():
        yield mock_redis

    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[get_redis] = override_get_redis

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as client:
        yield client

    app.dependency_overrides.clear()


@pytest.fixture
async def admin_user(test_session) -> User:
    """Create an admin user within the test transaction."""
    user = User(
        id=uuid.uuid4(),
        email="admin@test.com",
        hashed_password=hash_password("adminpass"),
        is_admin=True,
        is_active=True,
    )
    test_session.add(user)
    await test_session.flush()
    await test_session.refresh(user)
    return user


@pytest.fixture
async def regular_user(test_session) -> User:
    """Create a non-admin user within the test transaction."""
    user = User(
        id=uuid.uuid4(),
        email="user@test.com",
        hashed_password=hash_password("userpass"),
        is_admin=False,
        is_active=True,
    )
    test_session.add(user)
    await test_session.flush()
    await test_session.refresh(user)
    return user


@pytest.fixture
def admin_token(admin_user) -> str:
    """JWT access token for admin user."""
    token, _jti = create_access_token(str(admin_user.id), True)
    return token


@pytest.fixture
def user_token(regular_user) -> str:
    """JWT access token for regular (non-admin) user."""
    token, _jti = create_access_token(str(regular_user.id), False)
    return token


# ── Phase 2: File Ingestion Fixtures ──────────────────────────────────────


def make_upload_file(path: Path, filename: str | None = None) -> tuple:
    """Prepare a file for httpx multipart upload.

    Returns a tuple suitable for passing to httpx files= parameter.
    """
    name = filename or path.name
    content = path.read_bytes()
    return ("file", (name, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))


@pytest.fixture
def sample_xlsx_file(tmp_path) -> Path:
    """Create a valid .xlsx file with 5 rows of realistic contact data."""
    wb = Workbook()
    ws = wb.active
    ws.append(["First Name", "Last Name", "Email", "Company", "LinkedIn"])
    ws.append(["John", "Doe", "john@example.com", "Acme Corp", "linkedin.com/in/johndoe"])
    ws.append(["Jane", "Smith", "jane@smith.co", "Globex Inc", "linkedin.com/in/janesmith"])
    ws.append(["Bob", "Wilson", "bob@wilson.io", "Initech", "linkedin.com/in/bobwilson"])
    ws.append(["Alice", "Brown", "alice@brown.net", "Umbrella Corp", "linkedin.com/in/alicebrown"])
    ws.append(["Charlie", "Davis", "charlie@davis.org", "Wayne Enterprises", "linkedin.com/in/charliedavis"])
    path = tmp_path / "contacts.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def empty_xlsx_file(tmp_path) -> Path:
    """Create a .xlsx file with headers only, no data rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(["First Name", "Last Name", "Email"])
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def malformed_xlsx_file(tmp_path) -> Path:
    """Create a .xlsx file with no contact-identifiable data.

    Headers and data contain no names, emails, or LinkedIn URLs.
    Used for testing malformed row flagging.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Notes", "Random"])
    ws.append(["some random text", "12345"])
    ws.append(["another note here", "67890"])
    ws.append(["just a comment", "abcde"])
    path = tmp_path / "malformed.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def large_header_xlsx(tmp_path) -> Path:
    """Create a .xlsx file with various header name formats and mixed data.

    Tests header normalization and content sampling.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["fname", "lname", "E-mail", "Organization", "Phone Number", "domain"])
    ws.append(["John", "Doe", "john@example.com", "Acme", "+1-555-0100", "acme.com"])
    ws.append(["Jane", "Smith", "jane@smith.co", "Globex", "555-0101", "globex.com"])
    ws.append(["Bob", "Wilson", "bob@wilson.io", "Initech", "(555) 010-0102", "initech.com"])
    ws.append(["", "", "", "", "", ""])  # empty row (should be skipped)
    ws.append(["Alice", "Brown", "alice@brown.net", "Umbrella", "5550103", "umbrella.co"])
    ws.append(["Charlie", "Davis", "charlie@davis.org", "Wayne", "+44 20 7946 0958", "wayne.com"])
    ws.append(["Diana", "Evans", "diana@evans.io", "Stark", "555.010.0105", "stark.com"])
    ws.append(["Eve", "Foster", "eve@foster.com", "Queen", "+1 (555) 010-0106", "queen.com"])
    ws.append(["Frank", "Garcia", "frank@garcia.net", "Palmer", "555-0107", "palmer.com"])
    ws.append(["Grace", "Hill", "grace@hill.org", "Luthor", "+49 30 123456", "luthor.com"])
    ws.append(["Henry", "Ito", "henry@ito.co", "Osborn", "555-0109", "osborn.com"])
    path = tmp_path / "large_headers.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def csv_file(tmp_path) -> Path:
    """Create a .csv file for testing format rejection."""
    path = tmp_path / "data.csv"
    path.write_text("Name,Email\nJohn,john@example.com\n")
    return path


@pytest.fixture
def upload_dir_override(tmp_path, monkeypatch):
    """Override settings.upload_dir to use tmp_path during tests.

    Prevents disk writes to /data/uploads.
    """
    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir()
    monkeypatch.setattr(settings, "upload_dir", str(upload_dir))
    return upload_dir


# -- Phase 3: Enrichment Pipeline Fixtures ------------------------------------


@pytest.fixture
def mock_apollo_success_response():
    """Mock Apollo API success response with person match."""
    return {
        "person": {
            "id": "apollo-person-123",
            "first_name": "John",
            "last_name": "Doe",
            "email": "john@acme.com",
            "email_status": "verified",
            "linkedin_url": "linkedin.com/in/johndoe",
            "title": "VP Sales",
            "organization": {"name": "Acme Corp", "domain": "acme.com"},
        },
        "waterfall": {"status": "accepted"},
    }


@pytest.fixture
def mock_apollo_not_found_response():
    """Mock Apollo API response when no person match found."""
    return {"person": None}


@pytest.fixture
def mock_webhook_payload():
    """Mock Apollo webhook payload with phone data."""
    return {
        "request_id": "req-abc-123",
        "people": [
            {
                "id": "apollo-person-123",
                "waterfall": {
                    "phone_numbers": [
                        {
                            "raw_number": "+1-555-123-4567",
                            "sanitized_number": "+15551234567",
                            "confidence_cd": "high",
                            "status_cd": "valid_number",
                        }
                    ]
                },
            }
        ],
    }


@pytest.fixture
def sample_column_mappings():
    """Sample column mappings as stored in Job.column_mappings."""
    return [
        {"column": "Email", "detected_type": "email", "confidence": "HIGH"},
        {"column": "First Name", "detected_type": "first_name", "confidence": "HIGH"},
        {"column": "Last Name", "detected_type": "last_name", "confidence": "HIGH"},
        {"column": "Company", "detected_type": "company", "confidence": "MEDIUM"},
        {"column": "LinkedIn", "detected_type": "linkedin_url", "confidence": "HIGH"},
    ]
