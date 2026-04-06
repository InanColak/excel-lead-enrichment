import uuid
from pathlib import Path

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.jobs.detection import (
    COLUMN_TYPES,
    detect_column_types,
    get_contact_identifier_columns,
)
from app.jobs.models import Job, JobRow, JobStatus, RowStatus


def validate_upload(file: UploadFile) -> None:
    """Validate uploaded file format and content type.

    Raises HTTPException(400) for invalid format, HTTPException(413) for oversized files.
    """
    # Check file extension
    if file.filename is None or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are accepted. Please convert .xls or .csv files to .xlsx format.",
        )

    # Check content type — allow common Excel MIME types and octet-stream (browser default)
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }
    if file.content_type and file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid content type '{file.content_type}'. Expected an Excel .xlsx file.",
        )


async def check_file_size(file: UploadFile) -> bytes:
    """Read the file content and check size against max_upload_size_mb.

    Returns the file content bytes if within limits.
    Raises HTTPException(413) if file exceeds size limit.
    """
    max_bytes = settings.max_upload_size_mb * 1024 * 1024
    content = await file.read()
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File size exceeds the {settings.max_upload_size_mb}MB limit.",
        )
    return content


def save_uploaded_file(job_id: uuid.UUID, content: bytes) -> str:
    """Save uploaded file content to disk at {upload_dir}/{job_id}/original.xlsx.

    Returns the file path as a string.
    """
    job_dir = Path(settings.upload_dir) / str(job_id)
    job_dir.mkdir(parents=True, exist_ok=True)
    file_path = job_dir / "original.xlsx"
    file_path.write_bytes(content)
    return str(file_path)


def parse_excel_file(file_path: str) -> tuple[list[str], list[dict]]:
    """Parse an Excel file using openpyxl in read_only mode.

    Reads the first sheet only. Returns (headers, rows) where each row is a dict
    keyed by column header name.

    Raises HTTPException(400) for empty files or files exceeding max rows.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The Excel file contains no sheets.",
            )

        rows_iter = ws.iter_rows()

        # Read header row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The Excel file is empty (no header row found).",
            )

        headers = []
        for cell in header_row:
            val = str(cell.value).strip() if cell.value is not None else ""
            headers.append(val)

        # Check we have at least one non-empty header
        if not any(h for h in headers):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The Excel file has no valid column headers.",
            )

        # Read data rows
        data_rows: list[dict] = []
        max_rows = settings.max_rows_per_file

        for row in rows_iter:
            # Extract cell values
            values = [cell.value for cell in row]

            # Skip completely empty rows (all cells None or empty string)
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
                continue

            if len(data_rows) >= max_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"File exceeds the maximum of {max_rows} rows.",
                )

            # Build dict keyed by header
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(values):
                    val = values[i]
                    # Convert to string for consistent JSONB storage, keep None as None
                    if val is not None:
                        row_dict[header] = str(val) if not isinstance(val, (int, float, bool)) else val
                    else:
                        row_dict[header] = None
                else:
                    row_dict[header] = None

            data_rows.append(row_dict)

        if len(data_rows) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The Excel file has no data rows (header only).",
            )

        return headers, data_rows
    finally:
        wb.close()


async def create_job_from_upload(
    db: AsyncSession, user_id: uuid.UUID, file: UploadFile
) -> Job:
    """Orchestrate the full upload flow: validate, save, parse, create records.

    Returns the created Job with status PENDING_CONFIRMATION.
    """
    # 1. Validate format and content type
    validate_upload(file)

    # 2. Read and check file size
    content = await check_file_size(file)

    # 3. Create Job record with status UPLOADING
    job = Job(
        user_id=user_id,
        filename=file.filename or "unknown.xlsx",
        file_path="",  # Will be updated after save
        status=JobStatus.UPLOADING.value,
    )
    db.add(job)
    await db.flush()  # Get the generated UUID

    # 4. Save file to disk
    file_path = save_uploaded_file(job.id, content)
    job.file_path = file_path

    # 5. Parse Excel file
    _headers, data_rows = parse_excel_file(file_path)

    # 6. Create JobRow records for each non-empty row
    valid_count = 0
    for idx, row_data in enumerate(data_rows):
        job_row = JobRow(
            job_id=job.id,
            row_index=idx,
            raw_data=row_data,
            status=RowStatus.PENDING.value,
        )
        db.add(job_row)
        valid_count += 1

    # 7. Update Job with counts and transition to PENDING_CONFIRMATION
    job.status = JobStatus.PENDING_CONFIRMATION.value
    job.total_rows = valid_count
    job.valid_rows = valid_count
    job.error_rows = 0

    await db.flush()

    return job


async def get_job_by_id(
    db: AsyncSession, job_id: uuid.UUID, user_id: uuid.UUID
) -> Job:
    """Load a job by ID, verifying it belongs to the requesting user.

    Raises HTTPException(404) if not found or wrong user.
    """
    result = await db.execute(select(Job).where(Job.id == job_id))
    job = result.scalar_one_or_none()

    if job is None or job.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found.",
        )

    return job


def _require_pending_confirmation(job: Job) -> None:
    """Raise 409 if job is not in PENDING_CONFIRMATION status."""
    if job.status != JobStatus.PENDING_CONFIRMATION.value:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Job is in '{job.status}' status. Expected 'pending_confirmation'.",
        )


async def get_column_mappings(
    db: AsyncSession, job_id: uuid.UUID, user_id: uuid.UUID
) -> list[dict]:
    """Detect or return cached column mappings for a job.

    On first call, runs auto-detection on sample rows and caches in job.column_mappings.
    Subsequent calls return cached mappings.

    Raises HTTPException(404) if job not found or wrong user.
    Raises HTTPException(409) if job is not in PENDING_CONFIRMATION status.
    """
    job = await get_job_by_id(db, job_id, user_id)
    _require_pending_confirmation(job)

    # Return cached mappings if already detected/overridden
    if job.column_mappings is not None:
        return job.column_mappings

    # Load sample rows for detection (first 20)
    result = await db.execute(
        select(JobRow)
        .where(JobRow.job_id == job_id)
        .order_by(JobRow.row_index)
        .limit(20)
    )
    sample_job_rows = result.scalars().all()

    if not sample_job_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Job has no rows to detect columns from.",
        )

    # Extract headers from first row's raw_data keys and build sample dicts
    headers = list(sample_job_rows[0].raw_data.keys())
    sample_rows = [row.raw_data for row in sample_job_rows]

    # Run detection
    mappings = detect_column_types(headers, sample_rows)

    # Cache in job.column_mappings
    job.column_mappings = mappings
    await db.flush()

    return mappings


async def override_column_mappings(
    db: AsyncSession,
    job_id: uuid.UUID,
    user_id: uuid.UUID,
    overrides: list,
) -> list[dict]:
    """Override column mappings for a job.

    Merges overrides with existing mappings — only specified columns are updated.

    Raises HTTPException(404) if job not found or wrong user.
    Raises HTTPException(409) if job is not in PENDING_CONFIRMATION status.
    Raises HTTPException(422) if any mapped_type is not in COLUMN_TYPES.
    """
    job = await get_job_by_id(db, job_id, user_id)
    _require_pending_confirmation(job)

    # Validate all mapped_type values
    for override in overrides:
        if override.mapped_type not in COLUMN_TYPES:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Invalid column type '{override.mapped_type}'. Must be one of: {COLUMN_TYPES}",
            )

    # Ensure we have base mappings to merge with
    if job.column_mappings is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No column mappings detected yet. Call GET /mappings first.",
        )

    # Build override lookup
    override_map = {o.column: o.mapped_type for o in overrides}

    # Merge: update only specified columns
    updated_mappings = []
    for mapping in job.column_mappings:
        if mapping["column"] in override_map:
            updated_mappings.append(
                {
                    "column": mapping["column"],
                    "detected_type": override_map[mapping["column"]],
                    "confidence": "HIGH",  # User override is highest confidence
                }
            )
        else:
            updated_mappings.append(mapping)

    job.column_mappings = updated_mappings
    await db.flush()

    return updated_mappings


async def confirm_job(
    db: AsyncSession, job_id: uuid.UUID, user_id: uuid.UUID
) -> Job:
    """Confirm a job's column mappings and flag malformed rows.

    Transitions job to CONFIRMED status. Rows with no contact identifiers
    are flagged as ERROR with a descriptive message.

    Raises HTTPException(404) if job not found or wrong user.
    Raises HTTPException(409) if job is not in PENDING_CONFIRMATION status.
    Raises HTTPException(400) if column_mappings not set.
    """
    job = await get_job_by_id(db, job_id, user_id)
    _require_pending_confirmation(job)

    if job.column_mappings is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Column mappings not set. Call GET /mappings first to detect columns.",
        )

    # Determine which columns are mapped to contact identifier types
    identifier_types = get_contact_identifier_columns()
    identifier_columns = [
        m["column"]
        for m in job.column_mappings
        if m["detected_type"] in identifier_types
    ]

    # Load all rows for this job
    result = await db.execute(
        select(JobRow)
        .where(JobRow.job_id == job_id)
        .order_by(JobRow.row_index)
    )
    all_rows = result.scalars().all()

    valid_count = 0
    error_count = 0

    for row in all_rows:
        # Check if any identifier column has a non-empty value
        has_identifier = False
        for col in identifier_columns:
            val = row.raw_data.get(col)
            if val is not None and str(val).strip():
                has_identifier = True
                break

        if not has_identifier:
            row.status = RowStatus.ERROR.value
            row.error_message = f"No contact identifiers found in row {row.row_index}"
            error_count += 1
        else:
            # Keep as PENDING for enrichment
            valid_count += 1

    # Update job status and counts
    job.status = JobStatus.CONFIRMED.value
    job.valid_rows = valid_count
    job.error_rows = error_count

    await db.flush()

    return job
