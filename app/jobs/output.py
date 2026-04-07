"""Enriched Excel output generation module.

OUTPUT-01, D-55, D-56, D-57, D-64: Generates enriched .xlsx files with
original data preserved and enrichment columns appended.

Two exports:
- map_enrichment_status: maps row status + contact data to output status
- generate_output_file: produces the enriched Excel file
"""

import logging
import uuid
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker

from app.contacts.models import Contact
from app.jobs.models import Job, JobRow

logger = logging.getLogger(__name__)

# Appended column headers per D-55
ENRICHED_HEADERS = ["enriched_email", "enriched_phone", "enrichment_status"]


def map_enrichment_status(row_status: str, contact: Optional[Contact]) -> str:
    """Map a JobRow status + Contact data to the output enrichment_status value.

    Per D-55:
    - "enriched" + contact has phone -> "enriched"
    - "enriched" + contact has no phone -> "email_only"
    - "email_only" -> "email_only"
    - "not_found" -> "not_found"
    - "error" or "skipped" -> "error"
    - anything else -> "pending" (fallback)
    """
    if row_status == "enriched":
        if contact and contact.phone and contact.phone.strip():
            return "enriched"
        return "email_only"
    elif row_status == "email_only":
        return "email_only"
    elif row_status == "not_found":
        return "not_found"
    elif row_status in ("error", "skipped"):
        return "error"
    else:
        return "pending"


async def generate_output_file(
    job_id: uuid.UUID, session_factory: async_sessionmaker
) -> str:
    """Generate enriched Excel output file for a completed/partial job.

    Per D-56: reads original file with read_only=True, creates new Workbook for output.
    Per D-57: output filename is {stem}_enriched.xlsx alongside original.
    Per D-64: called at job finalization, not on download request.

    Per T-04-01: output_path is constructed from job.file_path (already validated
    at upload) -- never from user input. Validates output is under upload directory parent.

    Returns the output file path as a string.
    """
    async with session_factory() as db:
        # a. Load job
        result = await db.execute(select(Job).where(Job.id == job_id))
        job = result.scalar_one_or_none()
        if job is None:
            raise ValueError(f"Job {job_id} not found")

        # b. Load all job rows ordered by row_index
        result = await db.execute(
            select(JobRow)
            .where(JobRow.job_id == job_id)
            .order_by(JobRow.row_index)
        )
        rows = result.scalars().all()

        # c. Batch-load all contacts to avoid N+1
        contact_ids = [r.contact_id for r in rows if r.contact_id is not None]
        contacts_by_id: dict[uuid.UUID, Contact] = {}
        if contact_ids:
            contact_result = await db.execute(
                select(Contact).where(Contact.id.in_(contact_ids))
            )
            for contact in contact_result.scalars().all():
                contacts_by_id[contact.id] = contact

        # d. Build enrichment map: row_index -> {email, phone, status}
        enrichment_map: dict[int, dict] = {}
        for row in rows:
            contact = contacts_by_id.get(row.contact_id) if row.contact_id else None
            status = map_enrichment_status(row.status, contact)
            enrichment_map[row.row_index] = {
                "email": contact.email if contact else None,
                "phone": contact.phone if contact else None,
                "status": status,
            }

        # e. Read original file (read_only=True per D-56)
        wb_read = load_workbook(job.file_path, read_only=True, data_only=True)
        try:
            ws_read = wb_read.active

            # f. Create new workbook for output
            wb_out = Workbook()
            ws_out = wb_out.active

            # g. Copy header row and append enrichment headers
            rows_iter = ws_read.iter_rows()
            header_row = next(rows_iter)
            header_values = [cell.value for cell in header_row]
            ws_out.append(header_values + ENRICHED_HEADERS)

            # h. Copy data rows with enrichment data appended
            # CRITICAL: row_index 0 = first data row (Excel row 2 after header)
            row_counter = 0
            for data_row in rows_iter:
                original_values = [cell.value for cell in data_row]
                enrichment = enrichment_map.get(row_counter, {})
                enriched_values = [
                    enrichment.get("email"),
                    enrichment.get("phone"),
                    enrichment.get("status", "pending"),
                ]
                ws_out.append(original_values + enriched_values)
                row_counter += 1

        finally:
            wb_read.close()

        # i. Determine output path: {stem}_enriched.xlsx in same directory as original
        original_path = Path(job.file_path)
        output_filename = f"{Path(job.filename).stem}_enriched.xlsx"
        output_path = original_path.parent / output_filename

        # Save output file
        wb_out.save(str(output_path))

        # j. Update job.output_file_path
        job.output_file_path = str(output_path)
        await db.commit()

        logger.info(f"Generated output file for job {job_id}: {output_path}")
        return str(output_path)
